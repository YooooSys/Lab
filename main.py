import pandas as pd
from customtkinter import *
import customtkinter
from database import collection, log_collection, ValueValidality, DataCorrector, Log, CopyDataFieldNo_ID, Search
from tkinter import *
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

#JSON file
with open("option_properties.json","r", encoding="utf-8") as file:
    option_ = json.load(file)

# Giao diện CustomTkinter
theme = option_["theme"]
customtkinter.set_appearance_mode(theme)
app = CTk()
app.geometry("1920x1080")
app.title("Quản lý sinh viên")

def MaximizeWindow() -> None: 
    app.state('zoomed') 

app.minsize(1300, 500)
app.after(1, MaximizeWindow)

# Color
class Theme:
    def __init__(self) -> None:
        global theme
        self.default_color = "#313338" if theme == "Dark" else "#f0f0f0"
        self.light_color = "#3f4248" if theme == "Dark" else "#ffffff"
        self.dark_color = "#1e1f22" if theme == "Dark" else "#e3e5e8"
        self.text_color = "#c3c6ca" if theme == "Dark" else "#57595d"
        self.red = "#da373c"
        self.blue = "#5865f2"

# Table of content
header_frame = CTkFrame(master=app, height=30, fg_color=Theme().light_color)
header_frame.pack(expand=True, pady=(80, 0), padx=20, fill="both")

table_frame = CTkScrollableFrame(master=app, height=1000, fg_color=Theme().default_color)
table_frame.pack(expand=True, pady=(0, 20), padx=20, fill="both")


add_window = None
edit_window = None 
delete_window = None
sort_window = None
search_window = None
options_window = None

sort_column = StringVar() 
sort_column.set("MSSV")
sort_option = StringVar()
sort_option.set("Ascending Sort")


def PrintTitle() -> None:
    headers: list = [
        ("MSSV", 0, 100), ("Họ đệm", 1, 180), ("Tên", 2, 85), ("Giới tính", 3, 90), 
        ("Lớp", 4, 90), ("Ngày sinh", 5, 120), ("Email", 6, 240), ("Số TC đã có", 7, 100), 
        ("Tổng học phí", 8, 180), ("Học phí đã đóng", 9, 180), ("Còn nợ", 10, 180), ("Ghi chú", 11, 290)
    ]
    for text, column, width in headers:
        header = CTkLabel(
            master=header_frame, text=text, width=width, height=30, 
            font=("Arial", 15, "bold")
        )
        if text == "MSSV":
            header.grid(row=0, column=column, padx=(5,1), pady=0, sticky="nsew")
        elif text == "Ghi chú":
            header.grid(row=0, column=column, padx=1, pady=0, sticky="nsew")
        else:
            header.grid(row=0, column=column, padx=1, pady=0, sticky="nsew")

def SelectRow(row) -> None:
    global selected_row
    if selected_row != row:
        selected_row = row
        UpdateRowColors()


def UpdateRowColors() -> None:
    global selected_row
    for widget in table_frame.winfo_children():
        # Kiểm tra nếu widget là CTkFrame và có thuộc tính 'row'
        if isinstance(widget, CTkFrame) and 'row' in widget.grid_info():
            row = int(widget.grid_info()['row'])
            if row % 2 == 1:
                widget.configure(fg_color=Theme().light_color if selected_row == row else "transparent")
selected_row = None
        
def PrintElement(data, row, highlight_list) -> None:
    global selected_row
    fields = [
        ("mssv", 0, 98), ("hodem", 1, 180), ("name", 2, 84), ("gender", 3, 94), 
        ("class", 4, 90), ("birth", 5, 118), ("email", 6, 240), ("owned_cert", 7, 104), 
        ("tuition", 8, 180), ("payed", 9, 180), ("debt", 10, 180), ("note", 11, 284)
    ]
    frame_bg = Theme().light_color if selected_row == row else "transparent"
    row_frame = CTkFrame(
        master=table_frame,
        fg_color=frame_bg,
    )
    row_frame.grid(row=row, padx=0, pady=0, sticky="nsew")
    
    for field, column, width in fields:
        label_text = str(data.get(field, ""))
        if field in ["tuition", "payed", "debt"]:
            label_text += " VNĐ"

        label_bg = Theme().dark_color if field in highlight_list else "transparent"

        label = CTkLabel(
            master=row_frame,
            text=label_text,
            width=width,
            height=35,
            fg_color=label_bg,
            text_color=Theme().text_color
        )
        label.grid(row=0, column=column, padx=1, pady=1, sticky="nsew")
        row_frame.columnconfigure(column, weight=1)

        label.bind("<Button-3>", lambda e, d=data, r=row: ShowContextMenu(e, d))
        label.bind("<Button-1>", lambda e, r=row: SelectRow(r))
        label.bind("<Button-3>", lambda e, r=row: SelectRow(r))

def RefreshTable(documents=None) -> None:
    if documents == None:
        documents = list(collection.find())
    # Xóa các widget hiện tại trong `table_frame` (nếu có)
    for widget in table_frame.winfo_children():
        widget.grid_forget()  # Chỉ ẩn các widget thay vì xóa
    PrintTitle()
    try:
        # Hiển thị dữ liệu lên giao diện
        for idx, document in enumerate(documents):
            if len(search_result) == 0:
                PrintElement(document, idx * 2 + 1, [])  # Hiển thị dòng dữ liệu
            else:
                PrintElement(document, idx * 2 + 1, search_result[idx])
    except Exception as e:
        print("Error: ", e)

notificate_msg_queue = []
notificate_frame = CTkFrame(master=app, height=0, width=300, fg_color=Theme().default_color, corner_radius=0)

class Animation:
    def __init__(self,mode: int = 1, Obj: str="notify", end_pos: int=0, start_pos: int=0, y: int = 0, x: int = 0, speed: int = 10, target_width: int=0) -> None:
        self.mode = mode # mode == 1 then Expand, else Collapse
        self.Obj = Obj # 
        self.end_pos = end_pos
        self.start_pos = start_pos #start_pos > end_pos for Expand_right func
        self.y = y
        self.x = x
        self.speed = speed # the more larger the more slower
        self.target_width = target_width

    def Expand_right(self):
        if self.Obj == "notify":
            global notificate_msg_queue
            queue = notificate_msg_queue

        frame = queue[-1 if self.mode == 1 else 0]
        real_width = frame.real_width
        frame.frame.place(x = self.end_pos + self.start_pos - real_width, y = self.y)
        frame.frame.configure(width=real_width)

        condition = (real_width < self.target_width) if self.mode == 1 else (real_width > 0)

        if condition:
            if self.Obj == "notify":
                queue[-1 if self.mode == 1 else 0].real_width += 10 if self.mode == 1 else -10

            app.after(self.speed, lambda: Animation(mode=self.mode,
                                                    Obj=self.Obj, 
                                                    start_pos=self.start_pos, 
                                                    y=self.y, 
                                                    speed=self.speed, 
                                                    target_width=self.target_width).Expand_right())
        elif self.mode != 1: 
            queue.pop(0)
            if queue == []:
                if self.Obj == "notify":
                    notificate_frame.configure(height=0)
                    notificate_frame.place(x=app.winfo_width(), y=app.winfo_height())

class NotificateMsg:
    def __init__(self, 
                notificate_msg = None, 
                notificate_msg_real_width = None,
                notificate_msg_y = None, 
                notificate_msg_target_width = 260) -> None:
        
        self.frame = notificate_msg
        self.real_width = notificate_msg_real_width
        self.target_width = notificate_msg_target_width
        self.y = notificate_msg_y
    
    
def NotificateDestroy():
    global notificate_frame, notificate_msg_queue

    notificate_msg:NotificateMsg = notificate_msg_queue[0]

    Animation(mode=0, Obj="notify", end_pos=10,start_pos=notificate_msg.target_width + 10, speed=10, y=notificate_msg.y, target_width=0).Expand_right()



def Notificate(msg: str) -> None:
    global notificate_msg_queue, notificate_frame

    notificate_msg_target_width = 260
    notificate_msg_height = 60
    y = 10 + notificate_frame.winfo_height() - notificate_msg_height
    space_between = 10
    speed = 10

    dynamic_width = max(notificate_msg_target_width, 20 + len(msg) * 10)

    notificate_frame.place(x=app.winfo_width() - 360, y=app.winfo_height() - notificate_frame.winfo_height() - 120)
    notificate_frame.configure(height=space_between + notificate_frame.winfo_height() + notificate_msg_height)
    
    notificate_msg_frame = CTkFrame(master=notificate_frame, fg_color=Theme().light_color, corner_radius=8, height=notificate_msg_height, width=0)
    notificate_msg_frame.place(x=10, y=y)

    label = CTkLabel(master=notificate_msg_frame, text=msg, text_color=Theme().text_color)
    label.place(x=15, y=15)

    notificate_msg_queue.append(NotificateMsg(notificate_msg_frame, 0, notificate_frame.winfo_height()))

    Animation(Obj="notify", 
              end_pos=10, 
              start_pos=dynamic_width + 10, 
              y=notificate_frame.winfo_height(), 
              speed=speed,
              target_width=dynamic_width).Expand_right()
    
    # Sau 5 giây, tự động đóng thông báo
    app.after(5000, NotificateDestroy)


context_menu = None
context_menu_height = 0
def ShowContextMenu(event, data) -> None:
    global context_menu
    
    if context_menu is not None:
        context_menu.destroy()

    # Tạo menu tùy chỉnh
    context_menu = CTkFrame(app, corner_radius=8, fg_color=Theme().light_color, width=120)

    # Lấy vị trí
    x = event.x_root - app.winfo_x()
    y = event.y_root - app.winfo_y()

    width_limit = app.winfo_x() + app.winfo_width()
    height_limit = app.winfo_y() + app.winfo_height()

    if event.x_root > width_limit - 120:
        x = app.winfo_width() - 120
    if event.y_root > height_limit - 80:
        y = app.winfo_height() - 80

    def Animation():
        global context_menu_height
        context_menu.configure(height=context_menu_height)
        context_menu.place(x=x, y=y + (30 - context_menu_height/2))

        if context_menu_height < 68:
            context_menu_height += 5
            app.after(10, Animation)
        else:
            context_menu_height = 0

    Animation()

    # Nút Chỉnh sửa
    edit_button = CTkButton(
        master=context_menu,
        text="Chỉnh sửa",
        command=lambda: [EditDataWindow(data), context_menu.destroy()],
        fg_color="transparent",
        hover_color=Theme().default_color,
        text_color=Theme().text_color,
        corner_radius=8,
        width=100
    )
    edit_button.place(x=10, y=5)

    # Nút Xóa
    delete_button = CTkButton(
        master=context_menu,
        text="Xóa",
        command=lambda: [DeleteDataWindow(data), context_menu.destroy()],
        fg_color="transparent",
        hover_color=Theme().default_color,
        text_color=Theme().text_color,
        corner_radius=8,
        width=100
    )
    delete_button.place(x=10, y=35)

    # Hàm đóng menu
    def CloseContextMenu(event):
        global context_menu
        if context_menu is not None:
            context_menu.destroy()
            context_menu = None
    app.bind("<Button-1>", CloseContextMenu)


def AddDataWindow() -> None:
    global add_window
    
    gender_=StringVar()
    gender_.set("Nam")

    if add_window and add_window.winfo_exists():
        add_window.focus()
        return

    add_window = CTkToplevel(app, fg_color=Theme().default_color)
    add_window.title("Thêm dữ liệu")
    add_window.geometry("700x310+660+400")
    add_window.maxsize(700, 310)
    add_window.minsize(700, 310)
    add_window.grab_set()

    def AddData():
        data = {
            "mssv": mssv_entry.get(),
            "hodem": hodem_entry.get(),
            "name": name_entry.get(),
            "gender": gender_display[gender_.get()],
            "class": class_entry.get(),
            "birth": birth_entry.get(),
            "email": email_entry.get(),
            "owned_cert": owned_cert_entry.get(),
            "tuition": tuition_entry.get(),
            "payed": payed_entry.get(),
            "note": note_entry.get(),
        }


        # Check if any field (except `debt`) is empty
        if not all(data[key] for key in data if key not in ["debt","note"]):
            error_label.configure(text="Vui lòng nhập đầy đủ thông tin!")
            return
        
        # Check for 
        tuition: str = data["tuition"]
        payed: str = data["payed"]
        name: str = data["name"]
        second_name: str = data["hodem"]
        email: str = data["email"]
        owned_cert: str = data["owned_cert"]
        id: str = data["mssv"]

        error_text = ValueValidality(id, name, second_name, email, owned_cert, tuition, payed)

        if error_text != "":
            error_label.configure(text=error_text)
            error_text = ""
            return
         
        # Add the calculated `debt` to the data dictionary
        data["debt"] = int(data["tuition"]) - int(data["payed"])

        DataCorrector(data)

        try:
            collection.insert_one(data)  # Add data to MongoDB
            Log(_id=data["_id"], msg="Thêm sinh viên mới", type="data_change", new_data=data) # Write Log
            add_window.destroy()

            Notificate("Dữ liệu đã được thêm ✅") #Show notificate box

        except Exception as e:
            Log(_id=data["_id"], msg=e, type="error")
            error_label.configure(text=f"Lỗi: {e}")

    def CreateEntry(label_text, row, column):
        label = CTkLabel(master=add_window, text=label_text, text_color=Theme().text_color)
        label.grid(row=row, column=column, padx=(20, 5), pady=5)
        entry = CTkEntry(master=add_window, width=200)
        entry.grid(row=row, column=column + 1, padx=5, pady=5)
        return entry

    mssv_entry = CreateEntry("MSSV:", 0, 0)
    hodem_entry = CreateEntry("Họ đệm:", 1, 0)
    name_entry = CreateEntry("Tên:", 2, 0)
    class_entry = CreateEntry("Lớp:", 4, 0)
    birth_entry = CreateEntry("Ngày sinh:", 5, 0)
    email_entry = CreateEntry("Email:", 0, 2)
    owned_cert_entry = CreateEntry("Số TC đã có:", 1, 2)
    tuition_entry = CreateEntry("Tổng học phí:", 2, 2)
    payed_entry = CreateEntry("Học phí đã đóng:", 3, 2)
    note_entry = CreateEntry("Ghi chú:", 4, 2)

    gender_display = {
        "Nam": "Nam",
        "Nữ": "Nữ"
    }

    label = CTkLabel(master=add_window,text="Giới tính:")
    label.grid(row=3,column=0, pady = 5, padx = 5)

    gender_combobox = CTkComboBox(master=add_window, values=list(gender_display.keys()), variable=gender_,width=200)
    gender_combobox.grid(row=3,column=1)


    add_button = CTkButton(master=add_window, text="Thêm", command=AddData)
    add_button.grid(row=11, column=1, padx=20, pady=10)

    cancel_button = CTkButton(master=add_window, text="Xong", command=add_window.destroy)
    cancel_button.grid(row=11, column=3, padx=5, pady=10)

    error_label = CTkLabel(master=add_window, text="", text_color=Theme().red)
    error_label.grid(row=12, column=1, padx=0, pady=0)

# Hàm mở cửa sổ chỉnh sửa dữ liệu
def EditDataWindow(data) -> None:
    global edit_window
    
    gender_=StringVar()
    gender_.set(data["gender"])

    if edit_window and edit_window.winfo_exists():
        edit_window.focus()
        return

    edit_window = CTkToplevel(app, fg_color=Theme().default_color)
    edit_window.title("Chỉnh sửa dữ liệu")
    edit_window.geometry("700x310+660+400")
    edit_window.maxsize(700, 310)
    edit_window.minsize(700, 310)
    edit_window.grab_set()

    def UpdateData():
        updated_data = {
            "mssv": mssv_entry.get(),
            "hodem": hodem_entry.get(),
            "name": name_entry.get(),
            "gender": gender_display[gender_.get()],
            "class": class_entry.get(),
            "birth": birth_entry.get(),
            "email": email_entry.get(),
            "owned_cert": owned_cert_entry.get(),
            "tuition": tuition_entry.get(),
            "payed": payed_entry.get(),
            "note": note_entry.get(),
        }

    
        # Check if any field (except `debt`) is empty
        if not all(updated_data[key] for key in updated_data if key not in ["debt","note"]):
            error_label.configure(text="Vui lòng nhập đầy đủ thông tin!")
            return

        tuition = updated_data["tuition"]
        payed = updated_data["payed"]
        name = updated_data["name"]
        second_name = updated_data["hodem"]
        email: str = updated_data["email"]
        owned_cert: str = updated_data["owned_cert"]
        id = updated_data["mssv"]
        _id = data["_id"]

        error_text = ValueValidality(id, name, second_name, email, owned_cert, tuition, payed, _id)

        if error_text != "":
            error_label.configure(text=error_text)
            error_text = ""
            return

        # Add the calculated `debt` to the updated_data dictionary
        updated_data["debt"] = int(updated_data["tuition"]) - int(updated_data["payed"])
        DataCorrector(updated_data)

        old_data = CopyDataFieldNo_ID(data)

        if old_data != updated_data:
            try:
                collection.update_one({"_id": data["_id"]}, {"$set": updated_data})
                Log(_id=data["_id"], msg="Cập nhật dữ liệu sinh viên", type="data_change", new_data=updated_data, old_data=old_data)
                
                document = list(collection.find())
                RefreshTable(document)

                app.after(500, lambda: Notificate("Dữ liệu đã được cập nhật ✅"))
                edit_window.destroy()

            except Exception as e:
                Log(_id=data["_id"], msg=e, type="error")
        else:
            error_label.configure(text="")
    def CreateEntry(label_text, initial_value, row, column):
        label = CTkLabel(master=edit_window, text=label_text,text_color=Theme().text_color)
        label.grid(row=row, column=column, padx=(20, 5), pady=5)
        entry = CTkEntry(master=edit_window, width=200)
        entry.insert(0, initial_value)
        entry.grid(row=row, column=column + 1, padx=5, pady=5)
        return entry
    
    gender_display = {
        "Nam": "Nam",
        "Nữ": "Nữ"
    }

    # Lấy dữ liệu người dùng
    mssv_entry = CreateEntry("MSSV:", data.get("mssv", ""), 0, 0)
    hodem_entry = CreateEntry("Họ đệm:", data.get("hodem", ""), 1, 0)
    name_entry = CreateEntry("Tên:", data.get("name", ""), 2, 0)
    class_entry = CreateEntry("Lớp:", data.get("class", ""), 4, 0)
    birth_entry = CreateEntry("Ngày sinh:", data.get("birth", ""), 5, 0)
    email_entry = CreateEntry("Email:", data.get("email", ""), 0, 2)
    owned_cert_entry = CreateEntry("Số TC đã có:", data.get("owned_cert", ""), 1, 2)
    tuition_entry = CreateEntry("Tổng học phí:", data.get("tuition", ""), 2, 2)
    payed_entry = CreateEntry("Học phí đã đóng:", data.get("payed", ""), 3, 2)
    note_entry = CreateEntry("Ghi chú:", data.get("note", ""), 4, 2)

    label = CTkLabel(master=edit_window,text="Giới tính:")
    label.grid(row=3,column=0, pady = 5, padx = 5)

    gender_combobox = CTkComboBox(master=edit_window, values=list(gender_display.keys()), variable=gender_, width=200)
    gender_combobox.grid(row=3,column=1)

    # Nút chức năng
    update_button = CTkButton(master=edit_window, text="Cập nhật", command=UpdateData)
    update_button.grid(row=11, column=1, padx=20, pady=10)

    cancel_button = CTkButton(master=edit_window, text="Xong", command=edit_window.destroy)
    cancel_button.grid(row=11, column=3, padx=5, pady=10)

    error_label = CTkLabel(master=edit_window, text="", text_color=Theme().red)
    error_label.grid(row=12, column=1, columnspan=2)

def DeleteDataWindow(data) -> None:
    global delete_window
    # Kiểm tra nếu cửa sổ con đã mở thì không tạo thêm
    if delete_window and delete_window.winfo_exists():
        delete_window.focus()
        return

    delete_window = CTkToplevel(app, fg_color=Theme().default_color)
    delete_window.title("Xóa dữ liệu")
    delete_window.geometry("400x100+660+400")
    delete_window.attributes('-topmost', True)

    old_data = CopyDataFieldNo_ID(data)
    def DeleteData():
        Log(_id=data["_id"], msg="Xóa dữ liệu", type="data_change", old_data=old_data)
        collection.delete_one(data)
        delete_window.destroy()

        Notificate("Xóa dữ liệu thành công ✅")

    label = CTkLabel(master=delete_window,text= "Bạn có chắc chắn muốn xóa người này?", anchor="center", text_color=Theme().text_color)
    label.grid(row=1,column=0, pady = 5, padx = 10)

    confirm_button = CTkButton(master=delete_window, text="OK", command=DeleteData)
    confirm_button.grid(row=2, column=0, padx=10, pady=10, sticky="e") 

    cancel_button = CTkButton(master=delete_window, text="Hủy", command=delete_window.destroy)
    cancel_button.grid(row=2, column=1, padx=10, pady=10, sticky="w")

def SortDataWindow() -> None:
    global sort_window
    if sort_window and sort_window.winfo_exists():
        sort_window.focus()
        return
    
    sort_window = CTkToplevel(app, fg_color=Theme().default_color)
    sort_window.title("Sắp xếp")
    sort_window.geometry("350x130+660+400")
    sort_window.attributes('-topmost', True)

    display_value = {
        "MSSV": "mssv",
        "Họ đệm": "hodem",
        "Tên": "name",
        "Lớp": "class",
        "Giới tính": "gender",
        "Ngày sinh": "birth",
        "Email": "email",
        "Số TC đã có": "owned_cert",
        "Tổng học phí": "tuition",
        "Học phí đã đóng": "payed",
        "Còn nợ": "debt",
        "Ghi chú": "note"
        }
    sort_option_display = {
        "Ascending Sort": 1,
        "Descending Sort": -1
    }
    label = CTkLabel(master=sort_window,text="Chọn thuộc tính cần sắp xếp", text_color=Theme().text_color)
    label.grid(row=1,column=0, pady = 5, padx = 10)
    sort_column_combobox = CTkComboBox(master=sort_window, values=list(display_value.keys()), variable=sort_column, text_color=Theme().text_color)
    sort_column_combobox.grid(row=2,column=0, pady = 5, padx = 10)

    sort_option_combobox = CTkComboBox(master=sort_window, values=list(sort_option_display.keys()), variable=sort_option, text_color=Theme().text_color)
    sort_option_combobox.grid(row=2,column=1, pady = 5, padx = 10)

    def sort_database() -> None:
        document = collection.find().sort({display_value[sort_column.get()]: sort_option_display[sort_option.get()]})
        RefreshTable(document)
        sort_window.destroy()

    confirm_button = CTkButton(master=sort_window, text="OK", command=sort_database)
    confirm_button.grid(row=3, column=0, padx=10, pady=10, sticky="e") 

    cancel_button = CTkButton(master=sort_window, text="Hủy", command=sort_window.destroy)
    cancel_button.grid(row=3, column=1, padx=10, pady=10, sticky="w")

search_result=[]
debounce_id = None
def SearchDataWindow() -> None:
    global search_window
    if search_window and search_window.winfo_exists():
        search_window.focus()
        return

    # Tạo cửa sổ tìm kiếm
    search_window = CTkToplevel(app, fg_color=Theme().default_color)
    search_window.title("Tìm kiếm")
    search_window.geometry("550x130+660+400")
    search_window.attributes('-topmost', True)

    # Label và Entry cho từ khóa
    label = CTkLabel(master=search_window, text="Nhập từ khóa cần tìm:", text_color=Theme().text_color)
    label.grid(row=0, column=0, pady=5, padx=10)

    entry = CTkEntry(master=search_window, width=240)
    entry.grid(row=1, column=0, pady=5, padx=(20, 5))

    # Các checkbox tùy chọn tìm kiếm
    match_case = BooleanVar()
    match_case_check = CTkCheckBox(master=search_window, text="Match Case", variable=match_case, text_color=Theme().text_color)
    match_case_check.grid(row=1, column=1, pady=5, padx=(20, 5))

    match_whole_word = BooleanVar()
    match_whole_word_check = CTkCheckBox(master=search_window, text="Match Whole Word", variable=match_whole_word,text_color=Theme().text_color)
    match_whole_word_check.grid(row=1, column=2, pady=5, padx=5)

    collection.create_index([("$**", "text")])

    # Hàm tìm kiếm dữ liệu
    def SearchData(event=None) -> None:
        search_data = Search(entry, match_case, match_whole_word)
        pipeline, temp, fields = search_data[0], search_data[1], search_data[2]

        documents = list(collection.aggregate(pipeline))

        global search_result
        search_result = []
        if len(temp) != 0:
            for doc in documents:
                ftemp = []
                for field in fields:
                    value = str(doc.get(field, ""))
                    if temp.lower() in value.lower():
                        ftemp.append(field)
                search_result.append(ftemp)
        RefreshTable(documents)
    def SearchDataDebounced(event=None):
        global debounce_id
        if debounce_id is not None:
            app.after_cancel(debounce_id)
        debounce_id = app.after(400, SearchData)


    # Gắn sự kiện tìm kiếm khi người dùng gõ vào ô nhập
    entry.bind("<KeyRelease>", SearchDataDebounced)

    # Gắn sự kiện để tìm kiếm khi thay đổi giá trị checkbox
    def UpdateSearch():
        SearchData()

    def ResetSearchResult():
        global search_result
        search_result = []

    # Khi thay đổi giá trị của checkbox
    match_case_check.configure(command=UpdateSearch)
    match_whole_word_check.configure(command=UpdateSearch)

    search_window.protocol('WM_DELETE_WINDOW', lambda: (ResetSearchResult(), RefreshTable(), search_window.destroy()))

def OptionsWindow() -> None:
    global options_window
    if options_window and options_window.winfo_exists():
        options_window.focus()
        return

    # Tạo cửa sổ mới
    options_window = CTkToplevel(app, fg_color=Theme().default_color)
    options_window.title("Tùy chọn")
    options_window.geometry("250x200+660+400")  # Adjusted window size
    options_window.attributes('-topmost', True)

    # Update the theme without refreshing all elements separately
    def apply_theme(theme_name):
        global theme
        theme = theme_name
        customtkinter.set_appearance_mode(theme_name)

        # Update theme in the settings file
        with open("option_properties.json", "w", encoding="utf-8") as file:
            json.dump(option_, file, indent=4, ensure_ascii=False)

        # Refresh the UI elements that depend on the theme
        header_frame.configure(fg_color=Theme().light_color if theme_name == "Light" else Theme().dark_color)
        table_frame.configure(fg_color=Theme().default_color)
        notificate_frame.configure(fg_color=Theme().default_color)
        options_window.configure(fg_color=Theme().default_color)
        label.configure(text_color=Theme().text_color)
        RefreshTable()
        LoadButtons()

    def theme_change_button(selected_theme):
        if option_["theme"] != selected_theme:
            option_["theme"] = selected_theme
            apply_theme(selected_theme)

    # Create buttons for theme switching
    dark_button = CTkButton(
        master=options_window,
        text="",
        width=30,
        hover_color="#313338",
        fg_color="#313338",
        corner_radius=20,
        border_width=2 if option_["theme"] == "Dark" else 0,
        border_color=Theme().blue,
        command=lambda: theme_change_button("Dark")
    )

    light_button = CTkButton(
        master=options_window,
        text="",
        width=30,
        hover_color="#f0f0f0",
        fg_color="#f0f0f0",
        corner_radius=20,
        border_width=2 if option_["theme"] == "Light" else 0,
        border_color=Theme().blue,
        command=lambda: theme_change_button("Light")
    )

    label = CTkLabel(master=options_window, text="Chế độ sáng/tối:", text_color=Theme().text_color)
    label.grid(row=0, column=0, padx=(10, 3), pady=5)

    dark_mode_var = BooleanVar(value=customtkinter.get_appearance_mode() == "Dark")

    def theme_change():
        # Lấy trạng thái từ biến và thay đổi chế độ
        if dark_mode_var.get():
            customtkinter.set_appearance_mode("Dark")
            option_["theme"] = "Dark"
        else:
            customtkinter.set_appearance_mode("Light")
            option_["theme"] = "Light"

        with open("option_properties.json", "w", encoding="utf-8") as file:
            json.dump(option_, file, indent=4, ensure_ascii=False)

    label = CTkLabel(master=options_window, text = "Chế độ sáng/tối:", text_color=Theme().text_color)
    label.grid(row=0, column=0, padx= (10,3), pady=5)

    dark_mode_check = CTkSwitch(
        master=options_window,
        variable=dark_mode_var,
        text="",
        command=theme_change,
    )
    dark_mode_check.grid(row=0, column=1, pady=5, padx=10)

    export_button = CTkButton(
        master=options_window,
        text="Export to Excel",
        command=ExportToExcel
    )
    export_button.grid(row=1, column=0, padx=10, pady=10)

    import_button = CTkButton(
        master=options_window,
        text="Import from Excel",
        command=ImportFromExcel
    )
    import_button.grid(row=2, column=0, padx=10, pady=10)

def ImportFromExcel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    
    if file_path:
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            Notificate(f"Data imported successfully")

            column_mapping = {
                "Mã số sinh viên": "mssv",
                "Họ và đệm": "hodem",
                "Tên": "name",
                "Giới tính": "gender",
                "Lớp": "class",
                "Ngày sinh": "birth",
                "Email": "email",
                "Số tín chỉ đã đạt": "owned_cert",
                "Tổng học phí": "tuition",
                "Học phí đã đóng": "payed",
                "Học phí còn nợ": "debt",
                "Ghi chú": "note",
            } #mã hóa các title

            df.rename(columns=column_mapping, inplace=True)
            
            df.fillna("", inplace=True)

            data_list = df.to_dict(orient='records')

            for data in data_list:
                if collection.find_one({"mssv": data["mssv"]}):
                    Notificate(f"Record with mssv {data['mssv']} already exists. Skipping insertion.")
                else:
                    DataCorrector(data)
                    collection.insert_one(data)
                    Notificate(f"Record with mssv {data['mssv']} inserted into MongoDB.")
            
        except Exception as e:
            Notificate(f"Error importing or inserting data: {e}")


def ExportToExcel():
    data = collection.find()
    data_list = list(data)

    if not data_list:
        Notificate("No data found in the collection.")
        return

    df = pd.DataFrame(data_list)

    if "_id" in df.columns:
        df.drop(columns=["_id"], inplace=True)
    column_mapping = {
        "mssv": "Mã số sinh viên", 
        "hodem": "Họ và đệm",
        "name": "Tên",
        "gender": "Giới tính",
        "class": "Lớp",
        "birth": "Ngày sinh",
        "email": "Email",
        "owned_cert": "Số tín chỉ đã đạt",
        "tuition": "Tổng học phí",
        "payed": "Học phí đã đóng",
        "debt": "Học phí còn nợ",
        "note": "Ghi chú",
    } #việt hóa các title

    df.rename(columns=column_mapping, inplace=True)

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        try:
            df.to_excel(file_path, index=False, engine='openpyxl')

            workbook = load_workbook(file_path)
            sheet = workbook.active

            center_columns = [
                "Mã số sinh viên", "Lớp", "Số tín chỉ đã đạt",
                "Tổng học phí", "Học phí đã đóng", "Học phí còn nợ",
                "Ghi chú", "Giới tính"
            ]

            for col in sheet.columns:
                col_letter = get_column_letter(col[0].column)
                if sheet[col_letter + "1"].value in center_columns:
                    for cell in col:
                        cell.alignment = Alignment(horizontal="center", vertical="center") #căn giữa

            for col in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value))) #thay đổi kích thước cột
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2

            workbook.save(file_path)

            Notificate(f"Data exported successfully")
        except Exception as e:
            Notificate(f"Error exporting data: {e}")

class CTkTooltip: #animation khi người dùng hover
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None): #hiển thị text khi người dùng hover
        if self.tooltip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tooltip_window = Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{x}+{y}")
        label = Label(
            self.tooltip_window,
            text=self.text,
            background="yellow",
            relief="solid",
            borderwidth=1,
            font=("Arial", 10, "normal")
        )
        label.pack(ipadx=5, ipady=3)

    def hide_tooltip(self, event=None): #tắt hộp thoại khi người dùng đưa con trỏ ra khỏi button
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

buttons_data = [
    {"image_path": r"template/add_student.png", "command": AddDataWindow, "x": 20, "size": (20, 20), "tooltip": "Thêm sinh viên mới"},
    {"image_path": r"template/sort.png", "command": SortDataWindow, "x": 75, "size": (20, 20), "tooltip": "Sắp xếp dữ liệu"},
    {"image_path": r"template/search.png", "command": SearchDataWindow, "x": 130, "size": (20, 20), "tooltip": "Tìm kiếm dữ liệu"},
    {"image_path": r"template/refresh.png", "command": lambda: RefreshTable(documents=list(collection.find())), "x": 185, "size": (20, 20), "tooltip": "Làm mới bảng"},
    {"image_path": r"template/option.png", "command": OptionsWindow, "x": 250, "size": (20, 20), "tooltip": "Tùy chọn"},
]

def LoadButtons():
    for button in buttons_data:
        image = PhotoImage(file=button["image_path"]).subsample(*button["size"])
        btn = CTkButton(
            master=app,
            text="",
            image=image,
            command=button["command"],
            width=50,
            height=50,
            fg_color="transparent",
            hover_color=Theme().light_color,
        )
        btn.place(x=button["x"], y=18)
        CTkTooltip(btn, button["tooltip"])

LoadButtons()
PrintTitle()
RefreshTable()

app.mainloop()
